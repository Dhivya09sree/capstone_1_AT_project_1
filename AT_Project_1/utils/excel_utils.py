from openpyxl import load_workbook

def get_data(file_path, sheet_name):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        test_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # Ensure both username and password are present
                test_data.append((row[0], row[1]))
            else:
                print(f"Skipping invalid row: {row}")

        return test_data
    except Exception as e:
        print(f"Error reading data from Excel: {e}")
        return None



# def write_data(file_path, sheet_name, row_num, col_num, data):
#     df = pd.read_excel(file_path, sheet_name=sheet_name)
#     df.iat[row_num - 2, col_num] = data  # Adjust row_num to zero-based index
#     df.to_excel(file_path, sheet_name=sheet_name, index=False)
